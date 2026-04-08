from ..config import *



from concurrent.futures import ThreadPoolExecutor, as_completed
import csv
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from pathlib import Path
import re
import shutil
import subprocess
import xlwings as xw


# Root directory
BASE_DIR = Path(r"D:\NEUP")

# Angles: folders whose names end with "deg"
ANGLE_SUFFIX = "deg"

# Subdirectory that contains conductivity data under each ANGLE folder
CONDUCTIVITY_SUBDIR = "Conductivity"

# Name of the input file in each AZIMUTH directory
INP_FILENAME = "Input.inp"

# Sample frequency (Hz)
SAMPLE_FREQ = 50000

# Order of radial positions (TAB files) to compile into each AZIMUTH.TAB
RADIUS_ORDER = [
    "r9o",
    "r85o",
    "r8o",
    "r7o",
    "r6o",
    "r5o",
    "r4o",
    "r3o",
    "r2o",
    "r1o",
    "r0o",  # or r0i or r0
    "r1i",
    "r2i",
    "r3i",
    "r4i",
    "r5i",
    "r6i",
    "r7i",
    "r8i",
    "r85i",
    "r9i",
]

# Azimuth folder names
AZIMUTHS = ["90", "67.5", "45", "22.5", "00"]

# Mapping from AZIMUTH string to top-left Excel cell for pasting
AZIMUTH_CELL_MAP = {
    "90": "B9",
    "67.5": "B56",
    "45": "B105",
    "22.5": "B152",
    "00": "B201",
}

def parse_radius_name_from_dat(dat_path: Path) -> str:
    """
    Given a DAT filepath, return its radius tag basename.
    Example: D:\...\r9o.DAT -> "r9o"
    """
    return dat_path.stem  # "r9o", "r85i", etc.



def radius_decimal_from_name(radius_name: str) -> float:
    """
    Convert rXo / rXi / rX into decimal r/R.

    Naming rule:
      - digits after 'r' are decimal digits of r/R
      - decimal point is placed after the first digit

    Examples:
        r0o  -> 0.0
        r1o  -> 0.1
        r9i  -> 0.9
        r85o -> 0.85

    NO negative sign even for 'i'.
    """
    name = radius_name.lower()

    if not name.startswith("r"):
        raise ValueError(f"Unexpected radius name: {radius_name}")

    core = name[1:]

    if core.endswith(("o", "i")):
        digits = core[:-1]
    else:
        digits = core

    if not digits.isdigit():
        raise ValueError(f"Could not parse digits in radius name: {radius_name}")

    denom = 10 ** len(digits)
    return int(digits) / denom


def find_run_dirs_with_dats(azimuth_dir: Path) -> list[Path]:
    """Return immediate child directories of `azimuth_dir` that contain one or more *.DAT files."""
    run_dirs = [p for p in azimuth_dir.iterdir() if p.is_dir() and any(p.glob("*.DAT"))]
    return sorted(run_dirs, key=lambda p: p.name)

def find_extra_dats_by_dir_datname_uniqueness(azimuth_dir: Path) -> dict[Path, list[Path]]:
    """
    Find .DAT files in subdirectories under `azimuth_dir` whose basename (stem)
    does NOT match any .DAT stem directly in `azimuth_dir`.

    Returns a mapping: {run_dir: [dat_paths_to_process_in_that_run_dir]}.

    Notes:
      - Root-level DATs in `azimuth_dir` are NOT included in this mapping.
      - Returned lists are sorted by filename for repeatability.
    """
    root_stems = {p.stem for p in azimuth_dir.glob("*.DAT")}

    extra: dict[Path, list[Path]] = {}
    for dat_path in azimuth_dir.rglob("*.DAT"):
        if dat_path.parent == azimuth_dir:
            continue
        if dat_path.stem in root_stems:
            continue
        extra.setdefault(dat_path.parent, []).append(dat_path)

    for run_dir, dats in extra.items():
        extra[run_dir] = sorted(dats, key=lambda p: p.name)

    return extra

def process_run_dir_with_midas(
    run_dir: Path,
    azimuth_dir: Path,
    distance_bias: float,
    dat_paths: list[Path] | None = None,
) -> None:
    """
    Runs MIDAS for selected *.DAT files inside `run_dir`.

    Correctness constraint:
      - `run_dir / Input.inp` is mutated per-DAT; therefore calls for the same `run_dir`
        must not overlap in time.
      - Parallelism is safe across distinct run directories.

    Behavior:
      - Ensures APAC/MIDAS executable exists in `run_dir` (copies from APAC_EXE if needed).
      - Ensures Input.inp exists in `run_dir` (copies from `azimuth_dir` as a template if needed).
      - Does NOT skip existing *.TAB outputs; they are overwritten by the MIDAS run (desired).
    """
    target_exe = run_dir / APAC_EXE.name
    if not target_exe.is_file():
        shutil.copy2(APAC_EXE, target_exe)

    inp_path = run_dir / INP_FILENAME
    if not inp_path.is_file():
        template_inp = azimuth_dir / INP_FILENAME
        if template_inp.is_file():
            shutil.copy2(template_inp, inp_path)
        else:
            print(f"        !! No {INP_FILENAME} in run dir or AZIMUTH template — skipping: {run_dir}")
            return

    if dat_paths is None:
        dat_files = sorted(run_dir.glob("*.DAT"), key=lambda p: p.name)
    else:
        dat_files = [Path(p) for p in dat_paths if Path(p).parent == run_dir]
        dat_files = sorted(dat_files, key=lambda p: p.name)

    if not dat_files:
        return

    for dat_path in dat_files:
        try:
            radius_name = parse_radius_name_from_dat(dat_path)
            r_over_R = radius_decimal_from_name(radius_name)
            sample_time = compute_sample_time(dat_path)
            filename_base = dat_path.stem

            edit_inp_file(
                inp_path,
                r_over_R=r_over_R,
                sample_time=sample_time,
                filename_base=filename_base,
                distance_bias=distance_bias,
            )

            run_midas_for_dat(dat_path, target_exe)

        except Exception as e:
            print(f"          !! ERROR processing {dat_path.name} in {run_dir}: {e}")

def compute_sample_time(dat_path: Path, fs: int = SAMPLE_FREQ) -> int:
    """
    Load DAT with np.loadtxt, compute sample time as len(data) / fs.

    Cast to int; this should be an integer for valid files.
    """
    data = np.loadtxt(dat_path)
    n = data.shape[0] if hasattr(data, "shape") else len(data)
    sample_time = n / fs
    # Force integer (you can change this to round or validate)
    return int(round(sample_time))


_R_DIST_RE = re.compile(r"^(r\d{2})\s*=\s*([-+]?[\d\.Ee+-]+)")

def edit_inp_file(
    inp_path: Path,
    r_over_R: float,
    sample_time: int,
    filename_base: str,
    distance_bias: float = 0.0,
):
    """
    Open .INP (text) file and update lines starting with:
        r/R=
        measuretime=
        Filename=
        rXX=   (apply distance_bias)

    Then overwrite the file.
    """
    lines = []

    with inp_path.open("r") as f:
        for line in f:
            stripped = line.lstrip()

            # r/R
            if stripped.startswith("r/R="):
                lines.append(f"r/R={r_over_R:.2f}\n")

            # measuretime
            elif stripped.startswith("measuretime="):
                lines.append(f"measuretime={sample_time}\n")

            # Filename
            elif stripped.startswith("Filename="):
                lines.append(f"Filename={filename_base}\n")

            # rXX distance entries
            else:
                m = _R_DIST_RE.match(stripped)
                if m:
                    key = m.group(1)           # e.g. r01
                    val = float(m.group(2))    # original value
                    new_val = val + distance_bias

                    # Preserve formatting consistency
                    lines.append(f"{key}={new_val:.3f}\n")
                else:
                    lines.append(line)

    with inp_path.open("w") as f:
        f.writelines(lines)


def run_midas_for_dat(dat_path: Path, apac_exe: Path):
    """
    Execute MIDAS/APAC for a given DAT file.

    We assume APAC.exe uses the local Input.inp in the current working directory
    and writes a .TAB file with the same basename as the DAT file.

    If APAC requires command-line args, update this function accordingly.
    """
    cwd = dat_path.parent
    if not apac_exe.is_file():
        raise FileNotFoundError(f"APAC executable not found at: {apac_exe}")

    # Run APAC in the AZIMUTH directory
    # If you need to pass the input file name, add it to the args.
    subprocess.run([str(apac_exe)], cwd=cwd, check=True)


_FILENAME_RE = re.compile(r"\bFilename\s*=\s*([^\s#;]+)", re.IGNORECASE)

def filename_order_from_inp(inp_path: Path) -> list[str]:
    """
    Extract Filename= entries from an Input.inp file, preserving order.

    Returns a list of DAT *stems* (e.g., ["r0o","r05o",..."]).

    Notes:
      - Many templates keep the full intended ordering in commented lines (leading '*').
        We still parse those lines for ordering.
      - The active (non-commented) Filename= line is also parsed.
    """
    stems: list[str] = []
    for line in inp_path.read_text(errors="ignore").splitlines():
        s = line.strip()
        if not s:
            continue

        # Keep commented Filename= lines for ordering; just strip the comment marker.
        if s.startswith("*"):
            s = s.lstrip("*").strip()
            if not s:
                continue

        m = _FILENAME_RE.search(s)
        if not m:
            continue
        token = m.group(1)
        stems.append(Path(token).stem)
    return stems

def compile_azimuth_tab(
    azimuth_dir: Path,
    azimuth_name: str,
    executable_name: str,
    extra_run_dirs: list[Path] | None = None,
    radius_order: list[str] | None = None,
):
    extra_run_dirs = extra_run_dirs or []
    radius_order = radius_order or RADIUS_ORDER

    out_path = azimuth_dir / f"{azimuth_name}_{executable_name}.TAB"

    # Case-insensitive tab lookup
    def index_tabs(d: Path) -> dict[str, Path]:
        m = {}
        for p in d.glob("*.TAB"):
            m[p.stem.lower()] = p
        for p in d.glob("*.tab"):
            m[p.stem.lower()] = p
        return m

    root_tabs = index_tabs(azimuth_dir)
    extra_tabs = {}
    for rd in extra_run_dirs:
        extra_tabs.update(index_tabs(rd))

    chunks = []
    for stem in radius_order:
        key = stem.lower()
        tab_path = root_tabs.get(key) or extra_tabs.get(key)

        if tab_path and tab_path.exists():
            text = tab_path.read_text(errors="ignore").rstrip("\n")
            # Ensure each radius contributes at least one line
            chunks.append(text if text else "")
        else:
            # Placeholder row to preserve radial mapping in Excel
            chunks.append("")

    out_path.write_text("\n".join(chunks) + "\n")


def excel_range_from_top_left(top_left: str, nrows: int, ncols: int):
    """
    Utility to compute Excel-style range from top-left coordinate and size.
    E.g., ("B10", 3, 4) -> "B10:E12"
    """
    row, col = coordinate_to_tuple(top_left)  # (row_idx, col_idx)
    last_row = row + nrows - 1
    last_col_letter = get_column_letter(col + ncols - 1)
    first_col_letter = get_column_letter(col)
    return f"{first_col_letter}{row}:{last_col_letter}{last_row}"


_NUM_RE = re.compile(r"^[+-]?(\d+(\.\d*)?|\.\d+)([eE][+-]?\d+)?$")

def coerce_cell(value: str):
    """
    Convert a string cell to int/float if it looks numeric.
    Otherwise return None for blanks, or the original string.
    """
    if value is None:
        return None

    s = str(value).strip()
    if s == "":
        return None

    # Common tokens
    if s.lower() in {"nan", "na", "n/a"}:
        # Choose None or float('nan'); None is usually better for Excel readability
        return None

    # Fast numeric check (handles scientific notation)
    if _NUM_RE.match(s):
        # Prefer int if exact integer
        try:
            if "." not in s and "e" not in s.lower():
                return int(s)
            return float(s)
        except Exception:
            pass

    return s


def read_tab_as_2d_list_numeric(tab_path: Path):
    rows = []
    with tab_path.open("r", newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        for row in reader:
            rows.append([coerce_cell(c) for c in row])

    if not rows:
        return []

    max_cols = max((len(r) for r in rows), default=0)
    if max_cols == 0:
        return [[""] for _ in rows]

    # Pad to rectangle with None (blank cells)
    rect = [r + [None] * (max_cols - len(r)) for r in rows]
    return rect


def paste_tab_into_sheet(ws, azimuth: str, tab_path: Path, clear_pad=(0, 0)):
    top_left = AZIMUTH_CELL_MAP.get(azimuth)
    if not top_left:
        return

    data = read_tab_as_2d_list_numeric(tab_path)
    if not data:
        return

    nrows = len(data)
    ncols = len(data[0])

    start_row, start_col = coordinate_to_tuple(top_left)

    # Clear destination area first (optional but recommended)
    extra_r, extra_c = clear_pad
    for i in range(nrows + extra_r):
        for j in range(ncols + extra_c):
            ws.cell(row=start_row + i, column=start_col + j).value = None

    # Paste values (numbers will be real ints/floats)
    for i, row_vals in enumerate(data):
        for j, v in enumerate(row_vals):
            ws.cell(row=start_row + i, column=start_col + j).value = v


def phase1_run_midas_and_build_tabs(distance_bias: float, max_workers: int = 4) -> None:
    """
    PHASE 1:
      - Runs MIDAS in parallel across RUN DIRECTORIES (safe for Input.inp mutation)
      - Builds AZIMUTH TAB compilations in the order prescribed by Input.inp (Filename= lines)
      - Also processes eligible subfolder DATs when their DAT stem does not exist at the AZIMUTH root.

    Notes:
      - Existing *.TAB files are NOT skipped; they are overwritten (intended mass reprocess behavior).
    """
    print("=== PHASE 1: Running MIDAS + Building TAB files ===")

    executable_name = APAC_EXE.stem

    port_prefixes = ("P1_", "P2_", "P3_")
    azimuth_set = set(AZIMUTHS)

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        for angle_dir in BASE_DIR.iterdir():
            if not angle_dir.is_dir() or not angle_dir.name.endswith(ANGLE_SUFFIX):
                continue

            print(f"[ANGLE] {angle_dir.name}")

            cond_dir = angle_dir / CONDUCTIVITY_SUBDIR
            if not cond_dir.is_dir():
                print("    (no Conductivity directory, skipping)")
                continue

            for condition_dir in cond_dir.iterdir():
                if not condition_dir.is_dir():
                    continue

                print(f"  [CONDITION] {condition_dir.name}")

                for port_dir in condition_dir.iterdir():
                    if not port_dir.is_dir() or not port_dir.name.startswith(port_prefixes):
                        continue

                    for azimuth_dir in port_dir.iterdir():
                        if not azimuth_dir.is_dir() or azimuth_dir.name not in azimuth_set:
                            continue

                        azimuth_name = azimuth_dir.name

                        # Root run: azimuth_dir itself (process all root DATs, if any)
                        root_has_dat = next(azimuth_dir.glob("*.DAT"), None) is not None

                        # Extra DATs: basename uniqueness vs root DATs
                        extra_dats_by_dir = find_extra_dats_by_dir_datname_uniqueness(azimuth_dir)
                        extra_run_dirs = sorted(extra_dats_by_dir.keys(), key=lambda p: str(p).lower())

                        futures = []

                        if root_has_dat:
                            futures.append(
                                ex.submit(
                                    process_run_dir_with_midas,
                                    azimuth_dir,
                                    azimuth_dir,
                                    distance_bias,
                                    None,  # all root DATs
                                )
                            )

                        for run_dir in extra_run_dirs:
                            dats = extra_dats_by_dir[run_dir]
                            print(f"        [RUN] {run_dir.name} (extra: {run_dir.relative_to(azimuth_dir)})")
                            futures.append(
                                ex.submit(
                                    process_run_dir_with_midas,
                                    run_dir,
                                    azimuth_dir,
                                    distance_bias,
                                    dats,  # only unique extra DATs
                                )
                            )

                        # Wait for MIDAS work for this azimuth to finish before compiling
                        for f in as_completed(futures):
                            try:
                                f.result()
                            except Exception as e:
                                print(f"        !! MIDAS failed under {azimuth_dir}: {e}")

                        # Compile combined AZIMUTH TAB in prescribed order
                        try:
                            compile_azimuth_tab(
                                azimuth_dir,
                                azimuth_name,
                                executable_name,
                                extra_run_dirs=extra_run_dirs,
                            )
                        except Exception as e:
                            print(f"        !! ERROR compiling {azimuth_name}_{executable_name}.TAB: {e}")

    print("=== PHASE 1 COMPLETE ===")


def phase15_fix_r85_in_azimuth_tabs(make_backups: bool = True, tol: float = 1e-9):
    """
    Intermediate Phase:
    Edit existing {AZIMUTH}.TAB files so that if the first element of row 2 (r85o)
    or row 20 (r85i) is 8.5, change it to 0.85.

    Assumes compiled {AZIMUTH}.TAB has 21 rows in the prescribed order.
    """

    print("\n=== PHASE 1.5: Fixing r85* r/R values in compiled AZIMUTH TABs ===\n")

    n_tabs_found = 0
    n_tabs_modified = 0

    for angle_dir in BASE_DIR.iterdir():
        if not angle_dir.is_dir() or not angle_dir.name.endswith(ANGLE_SUFFIX):
            continue

        cond_root = angle_dir / CONDUCTIVITY_SUBDIR
        if not cond_root.is_dir():
            continue

        for condition_dir in cond_root.iterdir():
            if not condition_dir.is_dir():
                continue

            for port_dir in condition_dir.iterdir():
                if not port_dir.is_dir():
                    continue
                if not (port_dir.name.startswith("P1_") or
                        port_dir.name.startswith("P2_") or
                        port_dir.name.startswith("P3_")):
                    continue

                for az in AZIMUTHS:
                    azimuth_dir = port_dir / az
                    if not azimuth_dir.is_dir():
                        continue

                    compiled_tab = azimuth_dir / f"{az}.TAB"
                    if not compiled_tab.is_file():
                        continue

                    n_tabs_found += 1

                    # Read tab-delimited content
                    with compiled_tab.open("r", newline="") as f:
                        rows = list(csv.reader(f, delimiter="\t"))

                    if len(rows) < 21:
                        print(f"!! {compiled_tab} has {len(rows)} rows (<21). Skipping.")
                        continue

                    modified = False

                    def fix_row(idx: int):
                        nonlocal modified
                        if idx < 0:
                            idx = len(rows) + idx
                        if idx < 0 or idx >= len(rows):
                            return

                        if not rows[idx]:
                            return

                        # Ensure at least one column exists
                        first = rows[idx][0].strip() if len(rows[idx]) >= 1 else ""
                        if first == "":
                            return

                        try:
                            val = float(first)
                        except ValueError:
                            return

                        if abs(val - 8.5) <= tol:
                            rows[idx][0] = "0.85"
                            modified = True

                    # row 2 (index 1) = r85o; row 20 (index -2) = r85i
                    fix_row(1)
                    fix_row(-2)

                    if modified:
                        if make_backups:
                            bak = compiled_tab.with_suffix(compiled_tab.suffix + ".bak")
                            if not bak.exists():
                                compiled_tab.replace(bak)  # move original to .bak
                                # write new fixed file to original path
                                with compiled_tab.open("w", newline="") as f:
                                    w = csv.writer(f, delimiter="\t")
                                    w.writerows(rows)
                            else:
                                # backup already exists; overwrite in place
                                with compiled_tab.open("w", newline="") as f:
                                    w = csv.writer(f, delimiter="\t")
                                    w.writerows(rows)
                        else:
                            with compiled_tab.open("w", newline="") as f:
                                w = csv.writer(f, delimiter="\t")
                                w.writerows(rows)

                        n_tabs_modified += 1
                        print(f"Fixed: {compiled_tab}")

    print(f"\nFound compiled AZIMUTH TABs: {n_tabs_found}")
    print(f"Modified AZIMUTH TABs:       {n_tabs_modified}")
    print("\n=== PHASE 1.5 COMPLETE ===\n")


def copy_xlsm_to_modified(suffix):
    """
    Copy ALL .xlsm files under ANGLE folders into DESTINATION_DIR,
    appending '_mod' before the extension.
    """
    DESTINATION_DIR.mkdir(exist_ok=True)

    for angle_dir in BASE_DIR.iterdir():
        if not angle_dir.is_dir():
            continue
        if not angle_dir.name.endswith(ANGLE_SUFFIX):
            continue

        for xlsm_path in angle_dir.glob("*.xlsm"):
            dest_name = xlsm_path.stem + f"{suffix}.xlsm"
            dest_path = DESTINATION_DIR / dest_name
            shutil.copy2(xlsm_path, dest_path)

    print("Copied all .xlsm files into Modified with '_mod' suffix.")


def find_condition_and_port_for_xlsm(xlsm_path: Path, suffix):
    """
    Given an XLSM filename like '90deg_jf2.00_jg0.10_P1_KM4-11S.xlsm',
    infer the ANGLE, CONDITION, and PORT directory path.

    Returns:
        condition_dir (Path), port_dir (Path), angle_dir (Path)
    or (None, None, None) if not found.
    """
    name = xlsm_path.stem
    if name.endswith(suffix):
        name = name[:-len(suffix)]

    # Expect: {angle}_{jf}_{jg}_{port}_{probe}
    # Example: 90deg_jf2.00_jg0.10_P1_KM4-11S
    parts = name.split("_")
    if len(parts) < 5:
        print(f"Unexpected xlsm name format: {xlsm_path.name}")
        return None, None, None

    angle = parts[0]            # "90deg"
    jf_part = parts[1]          # "jf2.00"
    jg_part = parts[2]          # "jg0.10"
    port = parts[3]             # "P1"
    probe = parts[4]            # "KM4-11S"

    angle_dir = BASE_DIR / angle
    cond_dir = angle_dir / CONDUCTIVITY_SUBDIR
    if not cond_dir.is_dir():
        print(f"No Conductivity directory for angle: {angle_dir}")
        return None, None, None

    # Condition folder name should be "angle_jfX.XX_jgX.XX"
    condition_folder_name = f"{angle}_{jf_part}_{jg_part}"
    condition_dir = cond_dir / condition_folder_name
    if not condition_dir.is_dir():
        print(f"Condition dir not found: {condition_dir}")
        return None, None, None

    # PORT directory names are P1_..., P2_..., P3_..., trailing text identifies probe.
    # We'll try a few matching strategies.
    candidate_ports = [d for d in condition_dir.iterdir() if d.is_dir() and d.name.startswith(port + "_")]

    if not candidate_ports:
        print(f"No PORT dir starting with {port}_ under {condition_dir}")
        return None, None, None

    # Prefer exact match, then substring match, then first
    expected_dirname = f"{port}_{probe}"
    best_port_dir = next((d for d in candidate_ports if d.name == expected_dirname), None)
    if best_port_dir is None:
        best_port_dir = next((d for d in candidate_ports if probe in d.name), None)
    if best_port_dir is None:
        best_port_dir = candidate_ports[0]

    return condition_dir, best_port_dir, angle_dir


def infer_tab_shape_for_port(port_dir: Path, azimuths, default_rows=21, default_cols=50):
    """
    Look through existing {AZIMUTH}.TAB files under this PORT directory and return:
        (rows_to_clear, cols_to_clear)

    Uses the maximum rows/cols found among existing TABs.
    Falls back to defaults if none found.
    """
    max_rows = 0
    max_cols = 0

    for az in azimuths:
        tab_path = port_dir / az / f"{az}.TAB"
        if not tab_path.is_file():
            continue

        with tab_path.open("r", newline="") as f:
            reader = csv.reader(f, delimiter="\t")
            rows = list(reader)

        max_rows = max(max_rows, len(rows))
        max_cols = max(max_cols, max((len(r) for r in rows), default=0))

    if max_rows == 0:
        max_rows = default_rows
    if max_cols == 0:
        max_cols = default_cols

    return max_rows, max_cols


def clear_azimuth_block(ws, azimuth: str, nrows: int, ncols: int):
    top_left = AZIMUTH_CELL_MAP.get(azimuth)
    if not top_left:
        return

    start_row, start_col = coordinate_to_tuple(top_left)
    for i in range(nrows):
        for j in range(ncols):
            ws.cell(row=start_row + i, column=start_col + j).value = None


def find_existing_azimuth_tabs(port_dir: Path, azimuths, id):
    """
    Returns a dict {azimuth: tab_path} for existing {AZIMUTH}.TAB files.
    """
    tabs = {}
    for az in azimuths:
        tab_path = port_dir / az / f"{az}{id}.TAB"
        if tab_path.is_file():
            tabs[az] = tab_path
    return tabs


def paste_data_into_sheet(ws, azimuth: str, data_2d):
    """
    data_2d is a 2D list (already numeric-coerced) to paste starting at AZIMUTH_CELL_MAP[azimuth].
    """
    top_left = AZIMUTH_CELL_MAP.get(azimuth)
    if not top_left:
        return

    start_row, start_col = coordinate_to_tuple(top_left)

    for i, row_vals in enumerate(data_2d):
        for j, v in enumerate(row_vals):
            ws.cell(row=start_row + i, column=start_col + j).value = v


def phase2_edit_xlsm_from_tabs(suffix, id):
    """
    PHASE 2:
      Copies XLSM files into DESTINATION_DIR and populates them from AZIMUTH TAB files,
      with progress logs.
    """

    print("\n=== PHASE 2: Copying XLSM + Populating Sheets ===\n")

    # 1) Copy all .xlsm
    copy_xlsm_to_modified(suffix)

    # 2) Loop through copied XLSM files
    for xlsm_path in DESTINATION_DIR.glob("*.xlsm"):
        print(f"\n[XLSM] {xlsm_path.name}")

        condition_dir, port_dir, angle_dir = find_condition_and_port_for_xlsm(xlsm_path,suffix)

        if condition_dir is None or port_dir is None:
            print("  !! Could not resolve CONDITION/PORT — skipping workbook")
            continue

        print(f"  Matched:")
        print(f"    ANGLE    -> {angle_dir.name}")
        print(f"    CONDITION-> {condition_dir.name}")
        print(f"    PORT     -> {port_dir.name}")

        # Open workbook (preserve macros)
        try:
            wb = load_workbook(xlsm_path, keep_vba=True)
        except Exception as e:
            print(f"  !! ERROR opening workbook: {e}")
            continue

        ws = wb["2"]
        print("  Opened workbook")

        # Infer clear block size from any existing TABs under this PORT directory
        clear_rows, clear_cols = infer_tab_shape_for_port(
            port_dir=port_dir,
            azimuths=AZIMUTHS,
            default_rows=21,   # your invariant
            default_cols=50    # set to your known TAB col count if you have it
        )

        # Detect available tabs
        existing_tabs = find_existing_azimuth_tabs(port_dir, AZIMUTHS, id)

        # If exactly one TAB exists, use it as the donor for all azimuths
        donor_data = None
        donor_az = None
        if len(existing_tabs) == 1:
            donor_az, donor_tab_path = next(iter(existing_tabs.items()))
            print(f"  Only one TAB exists ({donor_az}); will paste it into ALL azimuth blocks.")
            donor_data = read_tab_as_2d_list_numeric(donor_tab_path)  # must return 2D list


        # 3) Loop over azimuths
        for azimuth in AZIMUTHS:
            # Always clear first
            print(f"    [AZIMUTH {azimuth}] Clearing {clear_rows}x{clear_cols} at {AZIMUTH_CELL_MAP.get(azimuth)}")
            clear_azimuth_block(ws, azimuth, clear_rows, clear_cols)

            # Single-TAB fallback: paste donor into every azimuth
            if donor_data is not None:
                print(f"    [AZIMUTH {azimuth}] Pasting donor TAB from {donor_az}")
                paste_data_into_sheet(ws, azimuth, donor_data)
                continue

            # Normal behavior (2+ tabs exist): paste this azimuth's own TAB if it exists
            tab_path = existing_tabs.get(azimuth)
            if tab_path is None:
                print(f"    [AZIMUTH {azimuth}] TAB missing — cleared only")
                continue

            print(f"    [AZIMUTH {azimuth}] Pasting {tab_path.name}")
            paste_tab_into_sheet(ws, azimuth, tab_path)  # your numeric paste function
            

        # Save workbook
        try:
            wb.save(xlsm_path)
            print("  Saved workbook")
        except Exception as e:
            print(f"  !! ERROR saving workbook: {e}")

    print("\n=== PHASE 2 COMPLETE ===\n")


def phase25_recalc_saveas_fallback(
    folder: Path,
    pattern="*.xlsm",
    visible=True,
    local_outdir: Path = Path(r"C:\Temp\neup_cachefix_out")
):
    folder = Path(folder)
    files = sorted(folder.glob(pattern))
    if not files:
        print(f"No files matched {pattern} in {folder}")
        return

    local_outdir.mkdir(parents=True, exist_ok=True)

    def open_any(xl, path_str: str):
        # normal
        try:
            wb = xl.Workbooks.Open(path_str, UpdateLinks=0, ReadOnly=False,
                                   IgnoreReadOnlyRecommended=True, AddToMru=False, CorruptLoad=0)
            return wb, "normal"
        except Exception:
            pass
        # protected view
        try:
            pv = xl.ProtectedViewWindows.Open(path_str)
            wb = pv.Edit()
            return wb, "protected_view"
        except Exception:
            pass
        # repair
        wb = xl.Workbooks.Open(path_str, UpdateLinks=0, ReadOnly=False,
                               IgnoreReadOnlyRecommended=True, AddToMru=False, CorruptLoad=1)
        return wb, "repair"

    print(f"\n=== PHASE 2.5 (SaveAs fallback): {len(files)} workbook(s) ===")
    print(f"Folder: {folder}")
    print(f"Local outdir: {local_outdir}\n")

    with xw.App(visible=visible, add_book=False) as app:
        app.display_alerts = False
        app.screen_updating = False
        xl = app.api

        for i, f in enumerate(files, start=1):
            src = Path(f)
            path_str = str(src)
            tmp = local_outdir / src.name

            print(f"[{i}/{len(files)}] {src.name}")

            wb = None
            mode = None
            try:
                wb, mode = open_any(xl, path_str)
                print(f"    Opened via: {mode}")

                # Recalc
                xl.CalculateFullRebuild()
                xl.Calculate()

                # Try normal Save first
                try:
                    wb.Save()
                    wb.Close(SaveChanges=False)
                    print("    Saved in place")
                    continue
                except Exception as e_save:
                    print(f"    (save failed) {e_save}")

                # SaveAs fallback to local writable path
                # FileFormat 52 = xlOpenXMLWorkbookMacroEnabled (.xlsm)
                try:
                    if tmp.exists():
                        tmp.unlink()

                    wb.SaveAs(str(tmp), FileFormat=52)
                    wb.Close(SaveChanges=False)

                    # Copy back over original
                    shutil.copy2(tmp, src)
                    print("    Saved via SaveAs -> copied back (cache should now exist)")

                except Exception as e_saveas:
                    print(f"    !! SaveAs fallback failed: {e_saveas}")
                    try:
                        wb.Close(SaveChanges=False)
                    except Exception:
                        pass

            except Exception as e_open:
                print(f"    !! OPEN ERROR: {e_open}")
                try:
                    if wb is not None:
                        wb.Close(SaveChanges=False)
                except Exception:
                    pass

    print("\n=== PHASE 2.5 COMPLETE ===\n")
